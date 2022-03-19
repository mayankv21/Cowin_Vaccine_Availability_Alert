from typing import List
import openpyxl
import datetime as dt

def generateAvailabilityExcel (finalOpList: List[dict], excelDumpPath:str, startDate:dt.datetime, endDate:dt.datetime):
    """ this function generate excel of available vaccines at centres with meta information like vaccine name, no of doses etc

    Args:
        finalOpList (List[dict]): list of dictionary(where each dictionary is object corresponding to centre)
        excelDumpPath (str): excel dump folder path from config.json
        startDate (dt.datetime): startDate
        endDate (dt.datetime): endDate
    """  
    

    startDatestr = dt.datetime.strftime(startDate, '%Y-%m-%d')
    endDatestr = dt.datetime.strftime(endDate, '%Y-%m-%d')


    wb = openpyxl.Workbook()
    sheet = wb.active
    currRowNo =2

    sheet.cell(column=1, row=1, value= "Date")
    sheet.cell(column=2, row=1, value= "District_name")
    sheet.cell(column=3, row=1, value= "block_name")
    sheet.cell(column=4, row=1, value= "Address")
    sheet.cell(column=5, row=1, value= "Fee_type")
    sheet.cell(column=6, row=1, value= "Available_capacity")
    sheet.cell(column=7, row=1, value= "Available_capacity_dose1")
    sheet.cell(column=8, row=1, value= "Available_capacity_dose2")
    sheet.cell(column=9, row=1, value= "Vaccine_name")
    sheet.cell(column=10, row=1, value= "Min_age_limit")
    sheet.cell(column=11, row=1, value= "Max_age_limit")

    for center in finalOpList:
        sheet.cell(column=1, row=currRowNo, value= center["date"])
        sheet.cell(column=2, row=currRowNo, value= center["district_name"])
        sheet.cell(column=3, row=currRowNo, value= center["block_name"])
        sheet.cell(column=4, row=currRowNo, value= center["address"])
        sheet.cell(column=5, row=currRowNo, value= center["fee_type"])
        sheet.cell(column=6, row=currRowNo, value= center["available_capacity"])
        sheet.cell(column=7, row=currRowNo, value= center["available_capacity_dose1"])
        sheet.cell(column=8, row=currRowNo, value= center["available_capacity_dose2"])
        sheet.cell(column=9, row=currRowNo, value= center["vaccine_name"])
        sheet.cell(column=10, row=currRowNo, value= center["min_age_limit"])
        sheet.cell(column=11, row=currRowNo, value= center["max_age_limit"])
        currRowNo = currRowNo+1


    genFileName = f"{startDatestr}_to_{endDatestr}.xlsx"
    wb.save(filename=f"{excelDumpPath}/{genFileName}")
        
